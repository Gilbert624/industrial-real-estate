"""
Report Generator Module
Provides PDF and Excel export functionality for the asset management system
"""

from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from models.database import DatabaseManager, Asset
import io

class ReportGenerator:
    """Generate PDF and Excel reports for assets, projects, and financials"""
    
    def __init__(self):
        self.db = DatabaseManager()
    
    # ==================== Helper Methods ====================
    
    def _get_all_assets(self):
        """Helper method to get all assets"""
        session = self.db.get_session()
        try:
            assets = session.query(Asset).order_by(Asset.name).all()
            return assets
        except Exception as e:
            print(f"Error getting all assets: {e}")
            return []
        finally:
            self.db.close_session(session)
    
    def _get_completion_percentage(self, project):
        """Helper method to calculate completion percentage from project"""
        if project.total_budget and project.total_budget > 0:
            actual_cost = float(project.actual_cost) if project.actual_cost else 0.0
            total_budget = float(project.total_budget)
            return int((actual_cost / total_budget) * 100)
        return 0
    
    # ==================== PDF Reports ====================
    
    def generate_portfolio_pdf(self, output_path=None):
        """
        Generate comprehensive portfolio report in PDF format
        
        Args:
            output_path: File path to save PDF. If None, returns BytesIO buffer
            
        Returns:
            File path or BytesIO buffer
        """
        try:
            # Create buffer or file
            if output_path:
                doc = SimpleDocTemplate(output_path, pagesize=letter)
            else:
                buffer = io.BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=letter)
            
            # Container for the 'Flowable' objects
            elements = []
            
            # Define styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1f4788'),
                spaceAfter=30,
                alignment=TA_CENTER
            )
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=16,
                textColor=colors.HexColor('#2c5aa0'),
                spaceAfter=12,
                spaceBefore=12
            )
            
            # Title
            title = Paragraph("Portfolio Report", title_style)
            elements.append(title)
            
            # Report date
            date_text = Paragraph(
                f"<para align=center>Generated on {datetime.now().strftime('%B %d, %Y')}</para>",
                styles['Normal']
            )
            elements.append(date_text)
            elements.append(Spacer(1, 0.3*inch))
            
            # ===== Executive Summary =====
            elements.append(Paragraph("Executive Summary", heading_style))
            
            # Get summary data
            try:
                assets = self._get_all_assets()
                projects = self.db.get_all_projects()
                cash_balance = self.db.get_cash_balance()
                active_projects = self.db.get_active_projects_count()
                
                total_asset_value = sum(float(a.current_valuation) for a in assets if a.current_valuation) if assets else 0
                total_projects_budget = self.db.get_total_projects_budget()
                
                summary_data = [
                    ['Metric', 'Value'],
                    ['Total Assets', str(len(assets))],
                    ['Total Asset Value', f'${total_asset_value:,.0f}'],
                    ['Active Projects', str(active_projects)],
                    ['Total Project Budget', f'${total_projects_budget:,.0f}'],
                    ['Current Cash Balance', f'${cash_balance:,.0f}'],
                ]
                
                summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c5aa0')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                
                elements.append(summary_table)
                elements.append(Spacer(1, 0.4*inch))
                
            except Exception as e:
                elements.append(Paragraph(f"Error loading summary: {str(e)}", styles['Normal']))
            
            # ===== Assets Section =====
            elements.append(Paragraph("Asset Portfolio", heading_style))
            
            try:
                if assets:
                    asset_data = [['Asset Name', 'Type', 'Region', 'Area (sqm)', 'Valuation']]
                    for asset in assets:
                        asset_type_str = asset.asset_type.value if hasattr(asset.asset_type, 'value') else str(asset.asset_type)
                        asset_data.append([
                            asset.name[:30],
                            asset_type_str,
                            asset.region,
                            f"{asset.land_area_sqm:,.0f}" if asset.land_area_sqm else 'N/A',
                            f"${float(asset.current_valuation)/1e6:.1f}M" if asset.current_valuation else 'N/A'
                        ])
                    
                    asset_table = Table(asset_data, colWidths=[2.2*inch, 1.3*inch, 1.2*inch, 1*inch, 1*inch])
                    asset_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a90e2')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 9),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ]))
                    
                    elements.append(asset_table)
                else:
                    elements.append(Paragraph("No assets found", styles['Normal']))
                    
            except Exception as e:
                elements.append(Paragraph(f"Error loading assets: {str(e)}", styles['Normal']))
            
            elements.append(PageBreak())
            
            # ===== Projects Section =====
            elements.append(Paragraph("Active Projects", heading_style))
            
            try:
                if projects:
                    project_data = [['Project Name', 'Status', 'Budget', 'Spent', 'Progress']]
                    for proj in projects:
                        status_str = proj.status.value if hasattr(proj.status, 'value') else str(proj.status)
                        budget_val = float(proj.total_budget) if proj.total_budget else 0
                        actual_cost_val = float(proj.actual_cost) if proj.actual_cost else 0
                        completion = self._get_completion_percentage(proj)
                        
                        project_data.append([
                            proj.project_name[:30],
                            status_str,
                            f"${budget_val/1e6:.1f}M",
                            f"${actual_cost_val/1e6:.1f}M",
                            f"{completion}%"
                        ])
                    
                    project_table = Table(project_data, colWidths=[2.5*inch, 1.2*inch, 1.2*inch, 1.2*inch, 0.8*inch])
                    project_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a90e2')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 9),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ]))
                    
                    elements.append(project_table)
                else:
                    elements.append(Paragraph("No projects found", styles['Normal']))
                    
            except Exception as e:
                elements.append(Paragraph(f"Error loading projects: {str(e)}", styles['Normal']))
            
            # Build PDF
            doc.build(elements)
            
            if output_path:
                return output_path
            else:
                buffer.seek(0)
                return buffer
                
        except Exception as e:
            # Return error message in a buffer if no output_path
            if output_path:
                raise e
            else:
                error_buffer = io.BytesIO()
                error_buffer.write(f"Error generating PDF: {str(e)}".encode())
                error_buffer.seek(0)
                return error_buffer
    
    # ==================== Excel Reports ====================
    
    def generate_financial_excel(self, output_path=None):
        """
        Generate comprehensive financial report in Excel format
        
        Args:
            output_path: File path to save Excel. If None, returns BytesIO buffer
            
        Returns:
            File path or BytesIO buffer
        """
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Define styles
            header_fill = PatternFill(start_color="2C5AA0", end_color="2C5AA0", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            center_align = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # ===== Summary Sheet =====
            ws_summary = wb.create_sheet("Summary")
            ws_summary['A1'] = "Financial Summary Report"
            ws_summary['A1'].font = Font(size=16, bold=True, color="1F4788")
            ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            
            # Summary metrics
            try:
                cash = self.db.get_cash_balance()
                now = datetime.now()
                monthly_income = self.db.get_monthly_income(now.year, now.month)
                monthly_expense = self.db.get_monthly_expense(now.year, now.month)
                
                ws_summary['A4'] = "Metric"
                ws_summary['B4'] = "Value"
                ws_summary['A4'].fill = header_fill
                ws_summary['A4'].font = header_font
                ws_summary['B4'].fill = header_fill
                ws_summary['B4'].font = header_font
                
                metrics = [
                    ("Cash Balance", f"${cash:,.2f}"),
                    ("This Month Income", f"${monthly_income:,.2f}"),
                    ("This Month Expense", f"${monthly_expense:,.2f}"),
                    ("Net Cash Flow", f"${monthly_income - monthly_expense:,.2f}"),
                ]
                
                row = 5
                for metric, value in metrics:
                    ws_summary[f'A{row}'] = metric
                    ws_summary[f'B{row}'] = value
                    ws_summary[f'B{row}'].alignment = Alignment(horizontal="right")
                    row += 1
                
                # Apply borders
                for row_cells in ws_summary['A4:B8']:
                    for cell in row_cells:
                        cell.border = border
                        
            except Exception as e:
                ws_summary['A6'] = f"Error: {str(e)}"
            
            # Column widths
            ws_summary.column_dimensions['A'].width = 25
            ws_summary.column_dimensions['B'].width = 20
            
            # ===== Assets Sheet =====
            ws_assets = wb.create_sheet("Assets")
            
            headers = ['Asset Name', 'Type', 'Region', 'Status', 'Land Area (sqm)', 'Valuation', 'Purchase Date']
            for col, header in enumerate(headers, 1):
                cell = ws_assets.cell(1, col, header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border
            
            try:
                assets = self._get_all_assets()
                for row, asset in enumerate(assets, 2):
                    asset_type_str = asset.asset_type.value if hasattr(asset.asset_type, 'value') else str(asset.asset_type)
                    status_str = asset.status.value if hasattr(asset.status, 'value') else str(asset.status)
                    
                    ws_assets.cell(row, 1, asset.name)
                    ws_assets.cell(row, 2, asset_type_str)
                    ws_assets.cell(row, 3, asset.region)
                    ws_assets.cell(row, 4, status_str)
                    ws_assets.cell(row, 5, asset.land_area_sqm if asset.land_area_sqm else 0)
                    ws_assets.cell(row, 6, float(asset.current_valuation) if asset.current_valuation else 0)
                    ws_assets.cell(row, 7, asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else '')
                    
                    # Format valuation as currency
                    ws_assets.cell(row, 6).number_format = '$#,##0'
                    
                    # Borders
                    for col in range(1, 8):
                        ws_assets.cell(row, col).border = border
                        
            except Exception as e:
                ws_assets['A2'] = f"Error loading assets: {str(e)}"
            
            # Column widths
            ws_assets.column_dimensions['A'].width = 30
            ws_assets.column_dimensions['B'].width = 20
            ws_assets.column_dimensions['C'].width = 18
            ws_assets.column_dimensions['D'].width = 15
            ws_assets.column_dimensions['E'].width = 15
            ws_assets.column_dimensions['F'].width = 18
            ws_assets.column_dimensions['G'].width = 18
            
            # ===== Transactions Sheet =====
            ws_trans = wb.create_sheet("Transactions")
            
            headers = ['Date', 'Type', 'Category', 'Amount', 'Asset', 'Description']
            for col, header in enumerate(headers, 1):
                cell = ws_trans.cell(1, col, header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border
            
            try:
                transactions = self.db.get_recent_transactions(100)  # Get more for Excel
                for row, tx in enumerate(transactions, 2):
                    tx_type_str = tx.transaction_type.value if hasattr(tx.transaction_type, 'value') else str(tx.transaction_type)
                    
                    ws_trans.cell(row, 1, tx.transaction_date.strftime('%Y-%m-%d'))
                    ws_trans.cell(row, 2, tx_type_str)
                    ws_trans.cell(row, 3, tx.category if tx.category else '')
                    ws_trans.cell(row, 4, abs(float(tx.amount)))
                    ws_trans.cell(row, 5, tx.asset.name if tx.asset else '')
                    ws_trans.cell(row, 6, tx.description[:50] if tx.description else '')
                    
                    # Format amount
                    ws_trans.cell(row, 4).number_format = '$#,##0.00'
                    
                    # Color code by type
                    if tx_type_str.lower() == "income":
                        ws_trans.cell(row, 2).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    else:
                        ws_trans.cell(row, 2).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    
                    # Borders
                    for col in range(1, 7):
                        ws_trans.cell(row, col).border = border
                        
            except Exception as e:
                ws_trans['A2'] = f"Error loading transactions: {str(e)}"
            
            # Column widths
            ws_trans.column_dimensions['A'].width = 12
            ws_trans.column_dimensions['B'].width = 12
            ws_trans.column_dimensions['C'].width = 18
            ws_trans.column_dimensions['D'].width = 15
            ws_trans.column_dimensions['E'].width = 25
            ws_trans.column_dimensions['F'].width = 40
            
            # ===== Projects Sheet =====
            ws_projects = wb.create_sheet("Projects")
            
            headers = ['Project Name', 'Status', 'Budget', 'Actual Cost', 'Variance', 'Completion %']
            for col, header in enumerate(headers, 1):
                cell = ws_projects.cell(1, col, header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border
            
            try:
                projects = self.db.get_all_projects()
                for row, proj in enumerate(projects, 2):
                    status_str = proj.status.value if hasattr(proj.status, 'value') else str(proj.status)
                    budget_val = float(proj.total_budget) if proj.total_budget else 0
                    actual_cost_val = float(proj.actual_cost) if proj.actual_cost else 0
                    variance = budget_val - actual_cost_val
                    completion = self._get_completion_percentage(proj)
                    
                    ws_projects.cell(row, 1, proj.project_name)
                    ws_projects.cell(row, 2, status_str)
                    ws_projects.cell(row, 3, budget_val)
                    ws_projects.cell(row, 4, actual_cost_val)
                    ws_projects.cell(row, 5, variance)
                    ws_projects.cell(row, 6, completion)
                    
                    # Format currency
                    for col in [3, 4, 5]:
                        ws_projects.cell(row, col).number_format = '$#,##0'
                    
                    # Format percentage
                    ws_projects.cell(row, 6).number_format = '0"%"'
                    
                    # Color code variance
                    if variance < 0:
                        ws_projects.cell(row, 5).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    else:
                        ws_projects.cell(row, 5).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    
                    # Borders
                    for col in range(1, 7):
                        ws_projects.cell(row, col).border = border
                        
            except Exception as e:
                ws_projects['A2'] = f"Error loading projects: {str(e)}"
            
            # Column widths
            ws_projects.column_dimensions['A'].width = 30
            ws_projects.column_dimensions['B'].width = 15
            ws_projects.column_dimensions['C'].width = 15
            ws_projects.column_dimensions['D'].width = 15
            ws_projects.column_dimensions['E'].width = 15
            ws_projects.column_dimensions['F'].width = 12
            
            # Save
            if output_path:
                wb.save(output_path)
                return output_path
            else:
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                return buffer
                
        except Exception as e:
            # Return error message in a buffer if no output_path
            if output_path:
                raise e
            else:
                error_buffer = io.BytesIO()
                error_buffer.write(f"Error generating Excel: {str(e)}".encode())
                error_buffer.seek(0)
                return error_buffer
